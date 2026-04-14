# ============================================================
# APP WEB - EXTRATO GENIAL | Design Stitch
# Redesign com paleta de cores do Stitch (Azul/Verde/Dourado)
# Rodar com: streamlit run app.py
# ============================================================

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
import unicodedata
import traceback

# ============================================================
# CONSTANTES
# ============================================================

TARGET_PAYOUT = "PAGAMENTO DE PAY OUT - PARCEIRO"
TARGET_PAYIN  = "RECEBIMENTO DE PAY IN DE PARCEIRO"
TARGET_BLOQUEIO = "BLOQUEIO - DEPOSITO JUDICIAL"
TARGET_DESBLOQUEIO = "DESBLOQUEIO JUDICIAL"
TARGETS = [TARGET_PAYOUT, TARGET_PAYIN, TARGET_BLOQUEIO, TARGET_DESBLOQUEIO]

# ============================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================

st.set_page_config(
    page_title="Extrato Genial",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================================
# SISTEMA DE DESIGN - STITCH DESIGN SYSTEM
# Paleta: Azul (#2563eb), Verde (#16a34a), Dourado (#d97706)
# ============================================================

DESIGN_TOKENS = """
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

        :root {
            /* ===== Paleta Principal - Azul ===== */
            --blue-950: #0a1628;
            --blue-900: #0f1f3a;
            --blue-800: #162d50;
            --blue-700: #1e3a5f;
            --blue-600: #2563eb;
            --blue-500: #3b82f6;
            --blue-400: #60a5fa;
            --blue-300: #93c5fd;
            --blue-200: #bfdbfe;
            --blue-100: #dbeafe;
            --blue-50: #eff6ff;

            /* ===== Paleta Secundária - Verde ===== */
            --green-900: #052e16;
            --green-800: #064e24;
            --green-700: #065f2e;
            --green-600: #16a34a;
            --green-500: #22c55e;
            --green-400: #4ade80;
            --green-300: #86efac;
            --green-200: #bbf7d0;
            --green-100: #dcfce7;
            --green-50: #f0fdf4;

            /* ===== Paleta Destaque - Dourado ===== */
            --gold-900: #451a03;
            --gold-800: #78350f;
            --gold-700: #92400e;
            --gold-600: #b45309;
            --gold-500: #d97706;
            --gold-400: #f59e0b;
            --gold-300: #fbbf24;
            --gold-200: #fcd34d;
            --gold-100: #fef3c7;
            --gold-50: #fffbeb;

            /* ===== Neutros ===== */
            --gray-900: #111827;
            --gray-800: #1f2937;
            --gray-700: #374151;
            --gray-600: #4b5563;
            --gray-500: #6b7280;
            --gray-400: #9ca3af;
            --gray-300: #d1d5db;
            --gray-200: #e5e7eb;
            --gray-100: #f3f4f6;
            --gray-50: #f9fafb;
            --white: #ffffff;

            /* Cores de Uso */
            --primary: var(--blue-600);
            --primary-dark: var(--blue-700);
            --primary-light: var(--blue-400);
            --secondary: var(--green-600);
            --secondary-dark: var(--green-700);
            --secondary-light: var(--green-400);
            --accent: var(--gold-500);

            --success: var(--green-600);
            --warning: var(--gold-500);
            --danger: #dc2626;

            --text-primary: var(--gray-900);
            --text-secondary: var(--gray-600);
            --text-muted: var(--gray-500);
            --bg-primary: var(--white);
            --bg-secondary: var(--gray-50);

            /* Sombras */
            --shadow-sm: 0 1px 2px rgba(0,0,0,0.05);
            --shadow: 0 1px 3px rgba(0,0,0,0.1);
            --shadow-md: 0 4px 6px -1px rgba(0,0,0,0.1);
            --shadow-lg: 0 10px 15px -3px rgba(0,0,0,0.1);
            --shadow-xl: 0 20px 25px -5px rgba(0,0,0,0.1);

            /* Bordas */
            --radius-sm: 6px;
            --radius: 10px;
            --radius-lg: 16px;
            --radius-xl: 24px;
            --radius-full: 9999px;

            /* Transições */
            --transition-fast: 150ms ease;
            --transition: 250ms ease;
        }

        * {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
        }

        /* Tabelas */
        .stDataFrame, .stDataFrame div, table {
            border: none !important;
        }

        th {
            border-bottom: 2px solid var(--gray-200) !important;
            background: var(--gray-50) !important;
        }

        td {
            border-bottom: 1px solid var(--gray-100) !important;
        }

        /* HEADER */
        .app-header {
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(12px);
            border-bottom: 1px solid var(--gray-200);
            padding: 1rem 2rem;
            box-shadow: var(--shadow-sm);
            display: flex;
            align-items: center;
            justify-content: space-between;
            border-radius: var(--radius-lg);
            margin-bottom: 2rem;
        }

        .app-header-brand {
            display: flex;
            align-items: center;
            gap: 1rem;
        }

        .app-header-icon {
            width: 48px;
            height: 48px;
            background: linear-gradient(135deg, var(--primary), var(--primary-dark));
            border-radius: var(--radius);
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 1.5rem;
            box-shadow: var(--shadow-sm);
        }

        .app-header-title {
            font-size: 1.5rem;
            font-weight: 800;
            color: var(--primary);
            margin: 0;
            letter-spacing: -0.5px;
        }

        .app-header-subtitle {
            font-size: 0.8rem;
            color: var(--text-secondary);
            font-weight: 500;
            margin: 0;
        }

        /* NAVEGAÇÃO */
        .nav-bar {
            display: flex;
            gap: 0.5rem;
            background: var(--gray-50);
            padding: 0.5rem;
            border-radius: var(--radius-full);
            margin-bottom: 2rem;
            border: 1px solid var(--gray-200);
        }

        .nav-item {
            padding: 0.75rem 1.5rem;
            border-radius: var(--radius-full);
            font-weight: 500;
            font-size: 0.9rem;
            color: var(--text-secondary);
            text-align: center;
            flex: 1;
            transition: all var(--transition-fast);
        }

        .nav-item:hover {
            background: var(--white);
            color: var(--primary);
        }

        .nav-item.active {
            background: var(--primary);
            color: var(--white);
            box-shadow: var(--shadow-md);
            font-weight: 600;
        }

        .nav-item.disabled {
            opacity: 0.5;
            cursor: not-allowed;
        }

        /* CARDS */
        .card {
            background: var(--white);
            border: 1px solid var(--gray-200);
            border-radius: var(--radius-lg);
            padding: 1.5rem;
            box-shadow: var(--shadow-sm);
            transition: all var(--transition);
            position: relative;
            overflow: hidden;
        }

        .card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
            background: linear-gradient(90deg, var(--primary), var(--green-500));
            opacity: 0;
            transition: opacity var(--transition);
        }

        .card:hover {
            box-shadow: var(--shadow-xl);
            transform: translateY(-8px);
            border-color: var(--gray-300);
        }

        .card:hover::before {
            opacity: 1;
        }

        /* SIGNATURE STYLE - Barra lateral */
        .card-signature::after {
            content: '';
            position: absolute;
            left: 0;
            top: 0;
            height: 100%;
            width: 4px;
            background: linear-gradient(180deg, var(--primary), var(--primary-light));
            border-radius: var(--radius-lg) 0 0 var(--radius-lg);
        }

        .card-signature.success::after {
            background: linear-gradient(180deg, var(--green-600), var(--green-400));
        }

        .card-signature.warning::after {
            background: linear-gradient(180deg, var(--gold-500), var(--gold-300));
        }

        .card-signature.error::after {
            background: linear-gradient(180deg, var(--danger), #ef4444);
        }

        .card-number {
            font-size: 2.5rem;
            font-weight: 800;
            line-height: 1;
            margin-bottom: 0.5rem;
            letter-spacing: -1px;
            background: linear-gradient(135deg, var(--primary), var(--green-600));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .card-label {
            font-size: 0.85rem;
            color: var(--text-secondary);
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        /* SEÇÕES */
        .section {
            background: linear-gradient(180deg, var(--gray-50) 0%, var(--gray-100) 100%);
            border-radius: var(--radius-xl);
            padding: 2rem;
            margin-bottom: 2rem;
            border: 1px solid var(--gray-200);
        }

        .section-title {
            font-size: 1.25rem;
            font-weight: 800;
            color: var(--text-primary);
            margin-bottom: 1.5rem;
            display: flex;
            align-items: center;
            gap: 0.75rem;
            letter-spacing: -0.3px;
        }

        /* UPLOAD BOX */
        .upload-zone {
            border: 2px dashed var(--gray-300);
            border-radius: var(--radius-lg);
            padding: 3rem 2rem;
            text-align: center;
            background: var(--white);
            transition: all var(--transition);
        }

        .upload-zone:hover {
            border-color: var(--primary);
            background: var(--blue-50);
            box-shadow: var(--shadow-md);
        }

        /* BOTÕES Streamlit */
        .stButton > button {
            background: linear-gradient(135deg, var(--primary), var(--primary-dark)) !important;
            color: var(--white) !important;
            border: none !important;
            border-radius: var(--radius-full) !important;
            padding: 0.875rem 2rem !important;
            font-size: 1rem !important;
            font-weight: 600 !important;
            transition: all var(--transition) !important;
            box-shadow: 0 4px 14px rgba(37, 99, 235, 0.4) !important;
        }

        .stButton > button:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 6px 20px rgba(37, 99, 235, 0.5) !important;
        }

        .stButton > button:disabled {
            opacity: 0.5 !important;
            cursor: not-allowed !important;
            transform: none !important;
        }

        /* STATUS BADGE */
        .status-badge {
            display: inline-flex;
            align-items: center;
            gap: 0.5rem;
            padding: 0.5rem 1rem;
            border-radius: var(--radius-full);
            font-size: 0.85rem;
            font-weight: 600;
            background: var(--white);
            box-shadow: var(--shadow-sm);
            border: 1px solid var(--gray-200);
        }

        .status-badge.success {
            color: var(--success);
            background: var(--green-50);
            border-color: var(--green-200);
        }

        /* DIVIDER */
        .divider {
            height: 1px;
            background: linear-gradient(90deg, transparent, var(--gray-300), transparent);
            margin: 2rem 0;
        }

        /* FOOTER */
        .app-footer {
            text-align: center;
            padding: 2rem 0;
            color: var(--text-muted);
            font-size: 0.8rem;
        }

        /* Esconder Streamlit UI */
        #MainMenu, footer, header {visibility: hidden;}

        /* Responsividade */
        @media (max-width: 768px) {
            .app-header { flex-direction: column; gap: 1rem; padding: 1rem; }
            .nav-bar { flex-direction: column; }
            .nav-item { padding: 0.75rem 1rem; }
            .card-number { font-size: 2rem; }
        }
    </style>
"""

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"\s+", " ", text)
    return text


_TARGETS_NORM = {_normalize_text(t) for t in TARGETS}


def formatar_excel(df_final):
    """Gera o arquivo Excel formatado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extrato Tratado"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    border       = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    center      = Alignment(horizontal="center", vertical="center")
    left_align  = Alignment(horizontal="left", vertical="center")

    headers    = ["Data", "HISTORICO", "Valor", "HISTORICO DE LANÇAMENTO"]
    col_widths = [15, 45, 18, 45]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22

    for idx, row in df_final.iterrows():
        r = idx + 2
        hist = row["HISTORICO"]

        dc = ws.cell(row=r, column=1, value=row["Data"])
        dc.number_format = "DD/MM/YYYY"
        dc.alignment = center
        dc.border = border

        hc = ws.cell(row=r, column=2, value=hist)
        hc.alignment = left_align
        hc.border = border
        hc.font = Font(name="Arial", size=10)

        vc = ws.cell(row=r, column=3, value=row["Valor"])
        vc.number_format = "#,##0.00"
        vc.alignment = center
        vc.border = border
        if hist in TARGETS:
            cor = "C00000" if row["Valor"] < 0 else "375623"
            vc.font = Font(name="Arial", color=cor, size=10, bold=True)
        else:
            vc.font = Font(name="Arial", size=10)

        hlc = ws.cell(row=r, column=4, value=row["HISTORICO DE LANÇAMENTO"])
        hlc.alignment = left_align
        hlc.border = border
        hlc.font = Font(name="Arial", size=10)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def processar_arquivo(uploaded_file):
    """Processa o arquivo Excel."""
    try:
        df = pd.read_excel(uploaded_file)

        colunas_obrigatorias = ["Data", "HISTORICO", "Valor", "HISTORICO DE LANÇAMENTO"]
        colunas_faltando = [col for col in colunas_obrigatorias if col not in df.columns]
        if colunas_faltando:
            return None, None, f"❌ Colunas faltando: {', '.join(colunas_faltando)}", None

        df["HISTORICO"] = df["HISTORICO"].astype(str).str.strip()
        df["HISTORICO_NORM"] = df["HISTORICO"].map(_normalize_text)

        df["Data"] = pd.to_datetime(df["Data"], errors="coerce").dt.date
        if df["Data"].isna().any():
            invalid = int(df["Data"].isna().sum())
            return None, None, f"❌ A coluna **Data** tem {invalid} valor(es) inválido(s).", None

        df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")
        if df["Valor"].isna().any():
            invalid = int(df["Valor"].isna().sum())
            return None, None, f"❌ A coluna **Valor** tem {invalid} valor(es) inválido(s).", None

        df_targets = df[df["HISTORICO_NORM"].isin(_TARGETS_NORM)].copy()
        df_others = df[~df["HISTORICO_NORM"].isin(_TARGETS_NORM)].copy()

        # Canonicaliza o histórico dos targets para manter consistência no arquivo final
        norm_to_canonical = {_normalize_text(t): t for t in TARGETS}
        df_targets["HISTORICO"] = df_targets["HISTORICO_NORM"].map(norm_to_canonical).fillna(df_targets["HISTORICO"])

        grouped = df_targets.groupby(["Data", "HISTORICO"], sort=False)["Valor"].sum().reset_index()
        grouped["HISTORICO DE LANÇAMENTO"] = grouped["HISTORICO"]

        df_others_clean = df_others[colunas_obrigatorias].copy()
        df_final = pd.concat([df_others_clean, grouped], ignore_index=True)
        # Ordenação ajuda leitura do “tratado” (mantém comportamento atual do app)
        df_final = df_final.sort_values(["Data", "HISTORICO"], kind="stable").reset_index(drop=True)

        total_payin = df_final[df_final["HISTORICO"] == TARGET_PAYIN]["Valor"].sum()
        total_payout = df_final[df_final["HISTORICO"] == TARGET_PAYOUT]["Valor"].sum()
        saldo = total_payin + total_payout

        stats = {
            "linhas_originais": len(df),
            "outros_mantidos": len(df_others_clean),
            "agrupados": len(grouped),
            "total_saida": len(df_final),
            "total_payin": total_payin,
            "total_payout": total_payout,
            "saldo": saldo,
        }

        return df_final, stats, None, None

    except Exception as e:
        debug = "".join(traceback.format_exception(type(e), e, e.__traceback__))
        return None, None, "❌ Erro ao processar o arquivo.", debug


# ============================================================
# ESTADO DA APLICAÇÃO
# ============================================================

if "current_page" not in st.session_state:
    st.session_state.current_page = "dashboard"
if "uploaded_file" not in st.session_state:
    st.session_state.uploaded_file = None
if "df_final" not in st.session_state:
    st.session_state.df_final = None
if "stats" not in st.session_state:
    st.session_state.stats = None
if "processing_done" not in st.session_state:
    st.session_state.processing_done = False

# ============================================================
# RENDERIZAÇÃO
# ============================================================

st.markdown(DESIGN_TOKENS, unsafe_allow_html=True)

# HEADER
st.markdown("""
    <div class="app-header">
        <div class="app-header-brand">
            <div class="app-header-icon">💎</div>
            <div>
                <p class="app-header-title">Extrato Genial</p>
                <p class="app-header-subtitle">Processamento Inteligente</p>
            </div>
        </div>
    </div>
""", unsafe_allow_html=True)

# NAVEGAÇÃO
col_nav1, col_nav2, col_nav3 = st.columns(3)
with col_nav1:
    if st.button("📊 Dashboard", use_container_width=True, type="primary" if st.session_state.current_page == "dashboard" else "secondary"):
        st.session_state.current_page = "dashboard"
        st.rerun()
with col_nav2:
    if st.button("⚡ Processamento", use_container_width=True, type="primary" if st.session_state.current_page == "processamento" else "secondary"):
        st.session_state.current_page = "processamento"
        st.rerun()
with col_nav3:
    st.button("📁 Histórico", use_container_width=True, disabled=True, type="secondary")

# ============================================================
# PÁGINA: DASHBOARD
# ============================================================

if st.session_state.current_page == "dashboard":
    st.markdown('<div class="section">', unsafe_allow_html=True)
    st.markdown('<div class="section-title"><span class="section-title-icon">📈</span> Visão Geral</div>', unsafe_allow_html=True)

    if st.session_state.processing_done and st.session_state.stats:
        stats = st.session_state.stats
        has_real_data = True
    else:
        stats = {"total_saida": 0, "total_payin": 0, "total_payout": 0, "saldo": 0}
        has_real_data = False

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown(f"""
            <div class="card card-signature">
                <div class="card-number">{stats['total_saida'] if has_real_data else "—"}</div>
                <div class="card-label">Transações</div>
            </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
            <div class="card card-signature success">
                <div class="card-number" style="font-size: 1.8rem;">
                    {f"R$ {stats['total_payin']:,.0f}" if has_real_data and stats['total_payin'] > 0 else "—"}
                </div>
                <div class="card-label">Pay In</div>
            </div>
        """, unsafe_allow_html=True)

    with col3:
        st.markdown(f"""
            <div class="card card-signature error">
                <div class="card-number" style="font-size: 1.8rem; background: linear-gradient(135deg, var(--danger), #ef4444); -webkit-background-clip: text; -webkit-text-fill-color: transparent;">
                    {f"R$ {abs(stats['total_payout']):,.0f}" if has_real_data and stats['total_payout'] < 0 else "—"}
                </div>
                <div class="card-label">Pay Out</div>
            </div>
        """, unsafe_allow_html=True)

    with col4:
        saldo_color = "var(--green-600)" if stats.get('saldo', 0) >= 0 else "var(--danger)"
        st.markdown(f"""
            <div class="card card-signature {'success' if stats.get('saldo', 0) >= 0 else 'error'}">
                <div class="card-number" style="font-size: 1.8rem; background: linear-gradient(135deg, {saldo_color}, {saldo_color}); -webkit-background-clip: text; -webkit-text-fill-color: transparent;">
                    {f"R$ {stats['saldo']:,.0f}" if has_real_data else "—"}
                </div>
                <div class="card-label">Saldo</div>
            </div>
        """, unsafe_allow_html=True)

    if not has_real_data:
        st.info("💡 **Nenhum dado ainda.** Vá para **Processamento** para fazer upload do extrato.")

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="section-title"><span class="section-title-icon">ℹ️</span> Como Funciona</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown('<div class="card"><div style="font-size: 2rem; margin-bottom: 0.5rem;">1️⃣</div><div style="font-weight: 700; margin-bottom: 0.5rem;">Upload</div><div style="color: var(--text-secondary); font-size: 0.9rem;">Faça upload do arquivo Excel (.xlsx)</div></div>', unsafe_allow_html=True)
    with col2:
        st.markdown('<div class="card"><div style="font-size: 2rem; margin-bottom: 0.5rem;">2️⃣</div><div style="font-weight: 700; margin-bottom: 0.5rem;">Processamento</div><div style="color: var(--text-secondary); font-size: 0.9rem;">Agrupamos PAY IN e PAY OUT automaticamente</div></div>', unsafe_allow_html=True)
    with col3:
        st.markdown('<div class="card"><div style="font-size: 2rem; margin-bottom: 0.5rem;">3️⃣</div><div style="font-weight: 700; margin-bottom: 0.5rem;">Download</div><div style="color: var(--text-secondary); font-size: 0.9rem;">Baixe o arquivo tratado e formatado</div></div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

# ============================================================
# PÁGINA: PROCESSAMENTO
# ============================================================

elif st.session_state.current_page == "processamento":
    st.markdown('<div class="section">', unsafe_allow_html=True)
    st.markdown('<div class="section-title"><span class="section-title-icon">⚡</span> Processamento de Extrato</div>', unsafe_allow_html=True)

    if not st.session_state.processing_done:
        # Estado 1: Upload
        st.markdown('<div class="upload-zone">', unsafe_allow_html=True)
        uploaded_file = st.file_uploader(
            "Selecione o arquivo Excel (.xlsx)",
            type=["xlsx"],
            help="Colunas: Data, HISTORICO, Valor, HISTORICO DE LANÇAMENTO",
            label_visibility="collapsed"
        )
        st.markdown('</div>', unsafe_allow_html=True)

        if uploaded_file:
            st.session_state.uploaded_file = uploaded_file
            st.success(f"✅ Arquivo: **{uploaded_file.name}**")

        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            processar_btn = st.button("⚡ PROCESSAR ARQUIVO", type="primary", disabled=uploaded_file is None, use_container_width=True)

        if processar_btn and uploaded_file:
            with st.spinner("⏳ Processando..."):
                df_final, stats, error, debug = processar_arquivo(uploaded_file)

            if error:
                st.error(error)
                if debug:
                    with st.expander("Detalhes técnicos (para suporte)"):
                        st.code(debug)
            else:
                st.session_state.df_final = df_final
                st.session_state.stats = stats
                st.session_state.processing_done = True
                st.rerun()

    else:
        # Estado 2: Sucesso
        stats = st.session_state.stats
        df_final = st.session_state.df_final

        st.markdown('<div class="status-badge success">✅ Processamento concluído!</div>', unsafe_allow_html=True)
        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # Cards de Estatística
        st.markdown("### 📊 Resumo")

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f'<div class="card card-signature"><div class="card-number">{stats["linhas_originais"]}</div><div class="card-label">Originais</div></div>', unsafe_allow_html=True)
        with col2:
            st.markdown(f'<div class="card card-signature success"><div class="card-number">{stats["agrupados"]}</div><div class="card-label">Agrupados</div></div>', unsafe_allow_html=True)
        with col3:
            st.markdown(f'<div class="card card-signature warning"><div class="card-number">{stats["outros_mantidos"]}</div><div class="card-label">Outros</div></div>', unsafe_allow_html=True)
        with col4:
            st.markdown(f'<div class="card card-signature"><div class="card-number">{stats["total_saida"]}</div><div class="card-label">Total</div></div>', unsafe_allow_html=True)

        # Métricas financeiras
        col5, col6, col7 = st.columns(3)
        with col5:
            st.markdown(f'<div class="card card-signature success"><div class="card-number" style="font-size: 1.8rem;">R$ {stats["total_payin"]:,.2f}</div><div class="card-label">Pay In</div></div>', unsafe_allow_html=True)
        with col6:
            st.markdown(f'<div class="card card-signature error"><div class="card-number" style="font-size: 1.8rem; background: linear-gradient(135deg, var(--danger), #ef4444); -webkit-background-clip: text; -webkit-text-fill-color: transparent;">R$ {abs(stats["total_payout"]):,.2f}</div><div class="card-label">Pay Out</div></div>', unsafe_allow_html=True)
        with col7:
            saldo_cls = 'success' if stats['saldo'] >= 0 else 'error'
            st.markdown(f'<div class="card card-signature {saldo_cls}"><div class="card-number" style="font-size: 1.8rem;">R$ {stats["saldo"]:,.2f}</div><div class="card-label">Saldo</div></div>', unsafe_allow_html=True)

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # Preview
        st.markdown("### 👀 Pré-visualização")
        st.dataframe(df_final.head(20), use_container_width=True, hide_index=True)
        if len(df_final) > 20:
            st.caption(f"Mostrando 20 de {len(df_final)} linhas")

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # Download
        st.markdown("### 📥 Download")
        excel_output = formatar_excel(df_final)
        nome_arquivo = st.session_state.uploaded_file.name.replace(".xlsx", "_tratado.xlsx")

        st.download_button(
            label="📥 BAIXAR EXCEL TRATADO",
            data=excel_output,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

        if st.button("🔄 NOVO ARQUIVO", type="secondary", use_container_width=True):
            st.session_state.processing_done = False
            st.session_state.uploaded_file = None
            st.session_state.df_final = None
            st.session_state.stats = None
            st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ============================================================
# FOOTER
# ============================================================

st.markdown("""
    <div class="app-footer">
        <div class="divider"></div>
        <p>💎 Extrato Genial • Automatização Inteligente</p>
    </div>
""", unsafe_allow_html=True)
