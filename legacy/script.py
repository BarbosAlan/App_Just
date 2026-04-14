# ============================================================
# CONFIGURAÇÕES - Edite aqui o nome dos arquivos
# ============================================================
ARQUIVO_ENTRADA = "01 - JANEIRO A DEZEMBRO GENIAL 4374-5 - MARCELO .xlsx"       # <- Coloque aqui o nome do seu arquivo Excel de entrada
ARQUIVO_SAIDA   = "01 - JANEIRO A DEZEMBRO GENIAL 4374-5 - MARCELO _tratado.xlsx"  # <- Coloque aqui o nome do arquivo Excel de saída
# ============================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TARGET_PAYOUT = "PAGAMENTO DE PAY OUT - PARCEIRO"
TARGET_PAYIN  = "RECEBIMENTO DE PAY IN DE PARCEIRO"
TARGETS       = [TARGET_PAYOUT, TARGET_PAYIN]

def processar():
    print(f"Lendo arquivo: {ARQUIVO_ENTRADA}")
    df = pd.read_excel(ARQUIVO_ENTRADA)
    df["HISTORICO"] = df["HISTORICO"].str.strip()
    df["Data"]      = pd.to_datetime(df["Data"]).dt.date

    df_targets = df[df["HISTORICO"].isin(TARGETS)].copy()
    df_others  = df[~df["HISTORICO"].isin(TARGETS)].copy()

    grouped = df_targets.groupby(["Data", "HISTORICO"], sort=False)["Valor"].sum().reset_index()
    grouped["HISTORICO DE LANÇAMENTO"] = grouped["HISTORICO"]

    df_others_clean = df_others[["Data", "HISTORICO", "Valor", "HISTORICO DE LANÇAMENTO"]].copy()
    df_final = pd.concat([df_others_clean, grouped], ignore_index=True)
    df_final = df_final.sort_values(["Data", "HISTORICO"]).reset_index(drop=True)

    print(f"  Linhas originais  : {len(df)}")
    print(f"  Outros (mantidos) : {len(df_others_clean)}")
    print(f"  PAY IN/OUT agrupados por dia: {len(grouped)}")
    print(f"  Total no arquivo de saída   : {len(df_final)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Extrato Tratado"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    payin_fill   = PatternFill("solid", start_color="E2EFDA")
    payout_fill  = PatternFill("solid", start_color="FDECEA")
    border       = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    center      = Alignment(horizontal="center", vertical="center")
    left_align  = Alignment(horizontal="left",   vertical="center")

    headers    = ["Data", "HISTORICO", "Valor", "HISTORICO DE LANÇAMENTO"]
    col_widths = [15, 45, 18, 45]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell            = ws.cell(row=1, column=i, value=h)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center
        cell.border     = border
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22

    for idx, row in df_final.iterrows():
        r    = idx + 2
        hist = row["HISTORICO"]

        if hist == TARGET_PAYIN:
            fill = payin_fill
        elif hist == TARGET_PAYOUT:
            fill = payout_fill
        else:
            fill = None

        dc                = ws.cell(row=r, column=1, value=row["Data"])
        dc.number_format  = "DD/MM/YYYY"
        dc.alignment      = center
        dc.border         = border
        if fill: dc.fill  = fill

        hc           = ws.cell(row=r, column=2, value=hist)
        hc.alignment = left_align
        hc.border    = border
        hc.font      = Font(name="Arial", size=10)
        if fill: hc.fill = fill

        vc                = ws.cell(row=r, column=3, value=row["Valor"])
        vc.number_format  = "#,##0.00"
        vc.alignment      = center
        vc.border         = border
        if hist in TARGETS:
            cor       = "C00000" if row["Valor"] < 0 else "375623"
            vc.font   = Font(name="Arial", color=cor, size=10, bold=True)
        else:
            vc.font   = Font(name="Arial", size=10)
        if fill: vc.fill = fill

        hlc           = ws.cell(row=r, column=4, value=row["HISTORICO DE LANÇAMENTO"])
        hlc.alignment = left_align
        hlc.border    = border
        hlc.font      = Font(name="Arial", size=10)
        if fill: hlc.fill = fill

    ws.freeze_panes = "A2"
    wb.save(ARQUIVO_SAIDA)
    print(f"\nArquivo salvo: {ARQUIVO_SAIDA}")

if __name__ == "__main__":
    processar()